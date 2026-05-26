"""
BulkWave Pro - Sample Data Generator
Creates data/sample_contacts.xlsx with 20 demo contacts.
Run once:  python create_sample.py
"""

import os
import random

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("openpyxl not found. Run: pip install openpyxl")
    raise

FIRST_NAMES = [
    "Aarav", "Priya", "Rahul", "Sneha", "Vikram",
    "Ananya", "Rohan", "Kavya", "Arjun", "Meera",
    "Kiran", "Divya", "Suresh", "Pooja", "Amit",
    "Nisha", "Ravi", "Sunita", "Deepak", "Anjali",
]

LAST_NAMES = [
    "Sharma", "Patel", "Singh", "Kumar", "Gupta",
    "Verma", "Shah", "Mehta", "Joshi", "Nair",
]

CITIES = ["Mumbai", "Delhi", "Bengaluru", "Chennai", "Hyderabad", "Pune"]

COUNTRY_CODE = "91"  # India


def random_phone() -> str:
    """Generate a realistic-looking Indian mobile number."""
    prefix = random.choice(["98", "97", "96", "95", "94", "93", "87", "86", "85", "70"])
    rest = "".join([str(random.randint(0, 9)) for _ in range(8)])
    return f"{COUNTRY_CODE}{prefix}{rest}"


def main():
    os.makedirs("data", exist_ok=True)
    output = "data/sample_contacts.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Header row styling
    header_fill = PatternFill("solid", fgColor="25D366")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    headers = ["Name", "Phone", "City", "Email"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 30

    # Data rows
    used_phones = set()
    for row_idx, fname in enumerate(FIRST_NAMES, start=2):
        lname = random.choice(LAST_NAMES)
        full_name = f"{fname} {lname}"
        phone = random_phone()
        while phone in used_phones:
            phone = random_phone()
        used_phones.add(phone)
        city = random.choice(CITIES)
        email = f"{fname.lower()}.{lname.lower()}@example.com"

        ws.cell(row=row_idx, column=1, value=full_name)
        ws.cell(row=row_idx, column=2, value=phone)
        ws.cell(row=row_idx, column=3, value=city)
        ws.cell(row=row_idx, column=4, value=email)

    # Also add 3 intentionally invalid numbers to test validation
    for i, bad_phone in enumerate(["123", "INVALID", "00000"], start=len(FIRST_NAMES) + 2):
        ws.cell(row=i, column=1, value=f"Bad Contact {i - len(FIRST_NAMES)}")
        ws.cell(row=i, column=2, value=bad_phone)
        ws.cell(row=i, column=3, value="N/A")
        ws.cell(row=i, column=4, value="bad@test.com")

    wb.save(output)
    print(f"[OK] Sample contacts saved to: {output}")
    print(f"     Rows: {len(FIRST_NAMES) + 3} (including 3 invalid numbers for testing)")


if __name__ == "__main__":
    main()
